"""
engine.py — owns all interaction with DuckDB.
Knows nothing about HTTP, FastAPI, or the web. Its only job:
load CSV/xlsx file(s), describe their structure, and run read-only queries.
"""

import os
import re
import duckdb
import openpyxl
import pandas

# Keywords that indicate a write/destructive operation.
# Matched as whole words, outside of string literals, to reduce
# false positives/bypasses.
_FORBIDDEN_KEYWORDS = [
    "insert", "update", "delete", "drop", "alter",
    "create", "attach", "copy", "pragma", "install", "load"
]

_MAX_ROWS = 1000  # hard cap on rows returned by any query


def _sanitize_name(raw: str) -> str:
    lowered = raw.lower()
    sanitized = re.sub(r"[^a-z0-9_]", "_", lowered)
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    return sanitized


def _sanitize_table_name(file_path: str) -> str:
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    sanitized = _sanitize_name(base_name)
    if not sanitized or not sanitized[0].isalpha():
        sanitized = f"table_{sanitized}" if sanitized else "table_unnamed"
    return sanitized


def _strip_string_literals(sql: str) -> str:
    """
    Replaces the contents of single-quoted string literals with a
    neutral placeholder, so forbidden-keyword scanning doesn't false-
    positive on legitimate search terms inside quotes (e.g. '%update%'
    searching for the word "update" in data, not a SQL command).
    Does not attempt to handle escaped quotes inside literals — SQL
    string literals in DuckDB use '' (doubled single quote) to
    represent a literal quote, which this simple approach doesn't
    special-case, but that's a rare enough edge case to accept for now.
    """
    return re.sub(r"'[^']*'", "''", sql)


def _contains_forbidden_keyword(sql: str) -> str | None:
    """
    Checks for forbidden keywords as whole words, OUTSIDE of string
    literals (so a search term like '%update%' doesn't false-positive
    on the word "update" appearing inside quotes). Returns the matched
    keyword, or None if the query is clean.
    """
    sql_without_literals = _strip_string_literals(sql)
    for keyword in _FORBIDDEN_KEYWORDS:
        if re.search(rf"\b{keyword}\b", sql_without_literals, re.IGNORECASE):
            return keyword
    return None


class DataEngine:
    def __init__(self):
        # ':memory:' means the database lives only in RAM — nothing
        # is written to disk. When the process exits, it's gone.
        self.con = duckdb.connect(database=":memory:")
        self.table_names: list[str] = []

        # Cap DuckDB's own memory usage as a safeguard against a runaway
        # query consuming all available RAM. This doesn't stop a slow
        # query, but it prevents the worst-case failure mode.
        self.con.execute("SET memory_limit = '2GB'")

    def _register_table_name(self, table_name: str):
        if table_name in self.table_names:
            raise ValueError(
                f"Table name '{table_name}' is already in use. "
                f"Rename the file/sheet or choose a different table name."
            )
        self.table_names.append(table_name)

    def load_csv(self, file_path: str, table_name: str | None = None):
        if table_name is None:
            table_name = _sanitize_table_name(file_path)

        self._register_table_name(table_name)

        self.con.execute(
            f"CREATE TABLE {table_name} AS SELECT * FROM read_csv_auto(?)",
            [file_path],
        )

    def load_xlsx(self, file_path: str, sheets: list[str] | None = None):
        """
        Loads sheets from an xlsx file as separate DuckDB tables.
        Table names follow the pattern '{filename}_{sheetname}', sanitized.

        If 'sheets' is None (default), every sheet in the file is loaded —
        this preserves the original behavior. Pass a list of sheet names
        to load only specific sheets instead.

        Assumes row 1 of each sheet is the header row — messier headers
        (title rows, merged cells) are a known limitation, not handled here.
        """
        base_name = _sanitize_table_name(file_path)
        workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

        sheet_names_to_load = sheets if sheets is not None else workbook.sheetnames

        for sheet_name in sheet_names_to_load:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in '{file_path}'.")

            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue  # skip genuinely empty sheets

            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]
            records = [dict(zip(headers, row)) for row in data_rows]

            table_name = f"{base_name}_{_sanitize_name(sheet_name)}"
            self._register_table_name(table_name)

            # DuckDB can register a pandas DataFrame directly as a
            # queryable virtual table — plain Python lists aren't supported.
            df = pandas.DataFrame(records)
            self.con.register("temp_sheet_view", df)
            self.con.execute(f"CREATE TABLE {table_name} AS SELECT * FROM temp_sheet_view")
            self.con.unregister("temp_sheet_view")

        workbook.close()

    def get_schema(self) -> dict:
        if not self.table_names:
            raise RuntimeError("No tables loaded yet.")

        tables = []
        for table_name in self.table_names:
            columns = self.con.execute(f"DESCRIBE {table_name}").fetchall()
            sample = self.con.execute(
                f"SELECT * FROM {table_name} LIMIT 5"
            ).fetchall()
            column_names = [col[0] for col in columns]

            tables.append({
                "table_name": table_name,
                "columns": [
                    {"name": col[0], "type": col[1]} for col in columns
                ],
                "sample_rows": [
                    dict(zip(column_names, row)) for row in sample
                ],
            })

        return {"tables": tables}

    def validate_query(self, sql: str):
        """
        Checks a query for safety issues WITHOUT executing it:
        - must start with SELECT
        - must not contain forbidden keywords (as whole words, outside
          of string literals)
        - must be syntactically valid, checked via EXPLAIN (which plans
          the query without running it)
        Raises ValueError with a clear message if any check fails.
        """
        lowered = sql.strip().lower()

        if not lowered.startswith("select"):
            raise ValueError("Only SELECT queries are allowed.")

        forbidden = _contains_forbidden_keyword(sql)
        if forbidden:
            raise ValueError(f"Query contains forbidden keyword: '{forbidden}'")

        try:
            self.con.execute(f"EXPLAIN {sql}")
        except Exception as e:
            raise ValueError(f"Query failed validation (syntax error): {str(e)}")

    def run_query(self, sql: str) -> dict:
        """
        Validates, then executes a read-only SQL query. Returns results
        as a list of dicts, capped at _MAX_ROWS. Indicates in the
        response if results were truncated.
        """
        self.validate_query(sql)

        result = self.con.execute(sql)
        column_names = [desc[0] for desc in result.description]
        rows = result.fetchmany(_MAX_ROWS + 1)  # fetch one extra to detect truncation

        truncated = len(rows) > _MAX_ROWS
        if truncated:
            rows = rows[:_MAX_ROWS]

        return {
            "rows": [dict(zip(column_names, row)) for row in rows],
            "truncated": truncated,
            "row_limit": _MAX_ROWS,
        }